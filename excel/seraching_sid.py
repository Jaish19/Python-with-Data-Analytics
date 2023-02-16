from openpyxl import load_workbook

wb = load_workbook('sample.xlsx')
sh = wb.get_sheet_by_name('Sheet')

sid = 56


print (sh.max_row)
for row_index in range(sh.max_row):
	row_index = row_index+1
	if (sh.cell(row=row_index, column=1).value == sid):
		print(sh.cell(row=row_index, column=2).value)